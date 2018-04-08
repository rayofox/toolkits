# coding:utf-8
import os,sys,shutil
import pymysql 
import html
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import date,datetime
from copy import copy,deepcopy
from pprint import pprint
from functools import wraps
import time,zipfile

def timeit(f):
    @wraps(f)
    def wrapper_function(*args, **kwargs):
        t1 = time.time()
        res = f(*args, **kwargs)
        t2 = time.time()
        print('== TimeSpan[{0} : {1:.2f}s]'.format(f.__name__,t2-t1))
        return res
    return wrapper_function
"""
1. 从禅道搜集信息 
2. 信息安装要求格式整理
3. 根据Excel模板填写Excel文件
4. 打包目录
"""

class excelHelper() :
    '''
    '''
    _group_task_file = {
        '开发/C++' : '开发\\C++\\开发任务检查表 - C++.xlsx',
        '开发/Java' : '开发\\Java\\开发任务检查表 - java.xlsx',
        '测试' : '测试\\测试任务检查表.xlsx',
    }
    _group_task_sheet = {
        '开发/C++' : '开发任务检查表',
        '开发/Java' : '开发任务检查表',
            '测试' : '测试任务检查表',
            
    }
    
    _group_bug_file = {
        '开发/C++' : '开发\\C++\\bug修复检查表 - C++.xlsx',
        '开发/Java' : '开发\\Java\\bug修复检查表 - java.xlsx',
        '测试' : '测试\\bug验证检查表.xlsx',
    }
    _group_bug_sheet = {
        '开发/C++' : 'Bug修复检查表',
            '开发/Java' : 'Bug修复检查表',
            '测试' : 'Bug验证检查表',
            
    }
    _user_group = {}
    def __init__(self,template_file,save_dir,start_date,end_date) :
        '''
        '''
        self._tempfile = template_file
        self._save_dir = save_dir
        self._top_path = os.path.join(save_dir,'SQA统计表[{0}-{1}]'.format(start_date,end_date) )
        self._zip_path = os.path.join(save_dir,'SQA统计表[{0}-{1}]-{2}.zip'.format(start_date,end_date,datetime.now().strftime('%Y%m%d-%H%M%S')) )
        pass
    @timeit
    def scaffold(self):
        '''
        '''
        #
        print('==> ')
        
        #
        wb = load_workbook(filename=self._tempfile,read_only=True)
        ws = wb['统计目标']
        for row in ws.rows :
            if not row[0].value or row[0].value == '用户' : continue
            self._user_group[row[0].value] = row[1].value
            pass
        #pprint(self._user_group)
        wb.close()
        #
        #
        shutil.rmtree(self._top_path,True)
        os.makedirs(self._top_path)
        #
        for gf,gs in [(self._group_task_file,self._group_task_sheet),(self._group_bug_file,self._group_bug_sheet)]:
            for gp,filename in gf.items() :
                fpath = os.path.join(self._top_path,filename)
                dn = os.path.dirname(fpath)
                if not os.path.exists(dn) : os.makedirs(dn)
                shutil.copyfile(self._tempfile,fpath)
                #
                wb = load_workbook(filename=fpath)
                for idx,sheetname in enumerate(wb.sheetnames) :
                    if sheetname == gs[gp] :
                        wb.active = idx
                        ws0 = wb[sheetname]
                        for user,group in self._user_group.items() :
                            if group != gp : continue
                            
                            ws = wb.copy_worksheet(ws0)
                            #
                            ws.conditional_formatting = deepcopy(ws0.conditional_formatting)
                            ws.title = user 
                            pass
                        pass
                    #
                    wb.remove(wb[sheetname])
                #
                wb.save(fpath)
                wb.close()
                pass
            
        pass
    @timeit
    def writeWorkloads(self,workloads):
        '''
        '''
        for gp,filename in self._group_task_file.items() :
            fpath = os.path.join(self._top_path,filename)
            #
            wb = load_workbook(filename=fpath)
            for sheetname in wb.sheetnames :
                user_workloads = workloads.get(sheetname,None)
                if not user_workloads : continue 
                ws = wb[sheetname]
                for workload in user_workloads:
                    row = 1+workload[0]
                    for i in range(len(workload)) :
                        col = i+1
                        cell = ws.cell(row,col)
                        cell.value = workload[i] 
                    pass
                pass
            #
            wb.save(fpath)
            wb.close()
            pass        
        pass
    @timeit
    def writeBugs(self,bugs) :
        '''
        '''
        for gp,filename in self._group_bug_file.items() :
            fpath = os.path.join(self._top_path,filename)
            #
            wb = load_workbook(filename=fpath)
            for sheetname in wb.sheetnames :
                user_bugs = bugs.get(sheetname,None)
                if not user_bugs : continue 
                ws = wb[sheetname]
                for bug in user_bugs:
                    row = 1+bug[0]
                    for i in range(len(bug)) :
                        col = i+1
                        cell = ws.cell(row,col)
                        cell.value = bug[i] 
                    pass
                pass
            #
            wb.save(fpath)
            wb.close()
            pass        
        pass
    @timeit
    def makeZip(self) :
        '''
        '''
        zipFile = zipfile.ZipFile(self._zip_path,'w',zipfile.ZIP_DEFLATED)
        os.chdir(self._save_dir)
        for dirpath,dirnames,filenames in os.walk(self._top_path):
            dpath = os.path.join(*dirpath.split(os.path.sep)[len(self._top_path.split(os.path.sep))-1:] )
            for filename in filenames:
                zipFile.write(os.path.join(dpath,filename))
                pass
            pass
        zipFile.close()
    
class ztHelper() :
    '''
    '''
    _db_cfg = {
        'host' : '192.168.5.206',
        'user' : 'dparser',
        'password' : 'Dparser@905',
        'db' : 'zentao',
        'port' : 3306,
        'charset' : 'utf8',
    }
    def __init__(self):
        '''
        '''
        self._dbc = None
        pass
    
    def _connectDB(self) :
        '''
        '''
        try :
            self._dbc = pymysql.connect(**self._db_cfg) 
        except Exception as e:
            print('== Connect to DB Failed : {0}'.format(e))
            return False
        self._cur = self._dbc.cursor()
        return True
    
    def _closeDB(self) :
        '''
        '''
        if self._dbc and self._cur :
            self._cur.close()
            self._dbc.close()
        return True 
    
    def _executeDB(self,sql,params=None) :
        '''
        '''
        rc = self._connectDB()
        if not rc : return rc
        try :
            self._cur.execute(sql,params)
            self._dbc.commit()
            pass
        except Exception as e:
            print('Execute SQL({0}) with params ({1}) failed : {2}'.format(sql,params,e))
            return False
        return True
    
    def _queryDB(self,sql,params=None,format_dict=True) :
        '''
        '''
        self._executeDB(sql,params)
        data_list =  self._cur.fetchall()
        #return data_list
        headers   = [ row[0] for row in self._cur.description]
        res_list = [ dict(zip(headers,line)) for line in data_list]
        return res_list
    def collectDataFromDB(self,start_date,end_date) :
        '''
        '''
        self._startDate,self._endDate = start_date,end_date
        self.collectBaseInfo()
        self.collectTasks()
        self.collectWorkloads()
        self.collectActions()
        self.collectBugFixed()
        self.collectBugVerified()
        pass
    
    def collectBaseInfo(self):
        '''
        '''
        sql = "SELECT account,realname FROM zt_user"
        res = self._queryDB(sql)
        #pprint(res)
        self._acc2name = {}
        for d in res:
            self._acc2name[d['account']] = d['realname']
            pass
    @timeit
    def collectTasks(self) :
        '''
        '''
        sql = """
        SELECT 
        P.`name` 所属项目,
        M.`name` 所属模块,
        T.`id` 任务ID,
        T.`name` 任务名,
        T.assignedTo 指派给,
        T.assignedDate 指派日期,
        T.type 任务类型,
        T.`status` 任务状态,
        T.pri 优先级,
        T.estStarted 预计开始,
        T.realStarted 实际开始,
        T.deadline 截止日期,
        T.estimate 最初预计,
        T.consumed 总消耗,
        T.`left` 预计剩余,
        T.openedBy 创建人,
        T.openedDate 创建日期,
        T.finishedBy 完成人,
        T.finishedDate 完成日期,
        T.canceledBy 取消人,
        T.canceledBy 取消日期,
        T.closedBy 关闭人,
        T.closedDate 关闭日期,
        T.closedReason 关闭原因,
        T.lastEditedBy 最后编辑作者,
        T.lastEditedDate 最后编辑日期,
        1
        FROM zt_task  T
        LEFT JOIN zt_project P ON P.id = T.project
        LEFT JOIN zt_module M ON M.id = T.module 
        WHERE 1=1
        #AND T.id = 3497
        AND T.lastEditedDate >= %s
        """
        self._tasks = self._queryDB(sql,[self._startDate])
        #pprint(self._tasks)
        return self._tasks
    @timeit
    def collectWorkloads(self) :
        '''
        '''
        sql = """
        SELECT *
        FROM zt_taskestimate
        WHERE 1=1
        #AND task = 3400
        AND consumed > 0
        AND date BETWEEN %s AND %s
        """
        self._workloads = self._queryDB(sql,[self._startDate,self._endDate])
        return self._workloads
    @timeit
    def collectActions(self) :
        '''
        '''
        sql = """
        SELECT 
        A.*,
        H.field,
        H.old,
        H.new,
        H.diff,
        1
        FROM zt_action A
        LEFT JOIN zt_history H ON H.action = A.id  
        WHERE 1=1
        #AND A.objectID = 3469
        AND A.date BETWEEN %s AND %s
        AND A.objectType IN ("bug","task")
        ORDER BY A.id
        """
        self._actions = self._queryDB(sql,[self._startDate,self._endDate])
        #pprint(self._actions)
        task2Event = self._task2Event = {}
        for d in self._actions:
            event = []
            action = d['action']
            actor = self._acc2name[d['actor']]
            field = d['field']
            if 'activated' == action :
                event.append('由{0}激活'.format(actor))
                pass
            elif 'editestimate' == action and 'left' == field:
                event.append('由{0}修改了预计剩余工时({1} -> {2})'.format(actor,d['old'],d['new']))
                pass
            elif 'finished' == action and 'status' == field:
                event.append('由{0}完成'.format(actor))
                pass
            elif 'closed' == action and 'status' == field:
                event.append('由{0}关闭'.format(actor))
                pass
            elif 'pause' == action and 'status' == field:
                event.append('由{0}暂停'.format(actor))
                pass
            #
            if event :
                taskId = d['objectID']
                dt = d['date'].strftime('%Y%m%d')
                #
                if taskId not in task2Event.keys() : task2Event[taskId] = {}
                te = task2Event[taskId]
                if dt not in te.keys() : te[dt] = []
                te[dt].extend(event)
            pass
        #pprint(task2Event)
        return task2Event
    @timeit
    def collectBugFixed(self) :
        '''
        '''
        sql = """SELECT 
        A.actor,A.action,A.date,A.extra,A.`comment`,
        B.id bugId,P.name 项目,
        B.severity bug等级 ,B.status bug状态,
        B.confirmed,B.activatedCount,
        B.openedDate , B.openedBy, B.resolvedDate,B.resolvedBy,B.closedDate,B.closedBy,
        1
        FROM zt_action A 
        LEFT JOIN zt_bug B ON A.objectID = B.id 
        LEFT JOIN zt_project P ON P.id = B.project
        LEFT JOIN zt_module M ON M.id = B.module
        LEFT JOIN zt_product PD ON PD.id = B.product 
        LEFT JOIN zt_user U ON U.account = B.resolvedBy
        WHERE 1=1
        AND A.objectType='bug' #AND A.action = 'resolved'
        AND B.resolvedDate BETWEEN %s AND %s AND U.role = 'dev'
        """
        _bugsFixed = self._queryDB(sql,[self._startDate,self._endDate])
        #pprint(_bugsFixed)
        self._bugId2FixedInfo = {}
        self._user2BugFixed = {}
        for bug in _bugsFixed :
            bugId = bug['bugId']
            user = self._acc2name[bug['resolvedBy']]
            if user not in self._user2BugFixed.keys() :
                self._user2BugFixed[user] = {}
                pass
            userBugFixedList = self._user2BugFixed[user]
            if bugId not in self._bugId2FixedInfo.keys() :
                self._bugId2FixedInfo[bugId] = copy(bug)
                pass
            bugInfo = self._bugId2FixedInfo[bugId]
            if bugId not in self._user2BugFixed[user].keys() :
                self._user2BugFixed[user][bugId] = bugInfo
                pass
            if 'bugconfirmed'==bug['action'] :
                bugInfo['confirmedDate'] = bug['date']
                pass
            elif 'activated'==bug['action'] :
                if 'activatedDate' not in bugInfo.keys() : # only record the last activated date 
                    bugInfo['activatedDate'] = bug['date']
                    pass
                pass
            elif 'resolved'==bug['action'] :
                if 'resolvedComment' not in bugInfo.keys() : # only record the last resovled comment
                    bugInfo['resolvedComment'] = bug['comment']
                    pass
                pass
            pass
        #pprint(self._user2BugFixed)
        pass
    @timeit
    def collectBugVerified(self) :
        '''
        '''
        sql = """
        SELECT 
        A.id,A.actor,A.action,A.date,A.extra,A.`comment`,
        P.name 项目,B.id bugId,
        B.openedDate , B.openedBy, B.resolvedDate,B.resolvedBy,B.closedDate,B.closedBy,
        BB.`name` buildName,BB.date buildDate,
        1
        FROM zt_action A 
        LEFT JOIN zt_bug B ON A.objectID = B.id 
        LEFT JOIN zt_project P ON P.id = B.project
        LEFT JOIN zt_module M ON M.id = B.module
        LEFT JOIN zt_product PD ON PD.id = B.product 
        LEFT JOIN zt_user U ON U.account = B.resolvedBy
        LEFT JOIN zt_build BB ON BB.id = B.resolvedBuild
        WHERE 1=1
        AND A.objectType='bug' AND A.action IN ('closed','activated')
        AND A.date BETWEEN %s AND %s AND U.role = 'qa' 
        """
        _bugsVerified = self._queryDB(sql,[self._startDate,self._endDate])
        self._user2BugVerified = {}
        for bug in _bugsVerified :
            aid = bug['id']
            user = self._acc2name[bug['actor']]
            if user not in self._user2BugVerified.keys() :
                self._user2BugVerified[user] = {}
                pass
            userBugFixedList = self._user2BugVerified[user]
            if aid not in self._user2BugVerified[user].keys() :
                self._user2BugVerified[user][aid] = bug
                pass
            if 'activated'==bug['action'] :
                bug['verifiedDate'] = bug['date']
                pass
            elif 'closed'==bug['action'] :
                bug['verifiedDate'] = bug['date']
                pass
        pass
    @timeit
    def parseData(self):
        self._user2workloads = {}
        self._user2bugs = {}
        
        self.parseTaskData()
        self.parseBugFixedData()
        self.parseBugVerifiedData()
        pass
    def parseTaskData(self):
        '''
        '''
        #
        xls_task_headers = ['序号','项目','开发模块','任务ID','日期','执行时长','是否填写工作日志','进度','修复bugs数','备注']
        self._d = {}
        task_workloads = self._d['task_workloads'] = {}
        #
        d_tasks = { task['任务ID'] : task for task in self._tasks }
        for wl in self._workloads:
            task = d_tasks[wl['task']]
            #pprint(wl)
            #pprint(task)
            #
            taskId = wl['task']
            dt = wl['date'].strftime('%Y%m%d')
            key = '{0}-{1}'.format(taskId,dt)
            if wl['account'] not in task_workloads : task_workloads[wl['account']] = {}
            if key not in task_workloads[wl['account']].keys() : 
                task_workloads[wl['account']][key] = [0 for i in range(len(xls_task_headers))]
                row = task_workloads[wl['account']][key]
                row[0] = len(task_workloads[wl['account']]) 
                row[1] = html.unescape(task['所属项目'])
                row[2] = html.unescape(task['任务名'])
                row[3] = (task['任务ID'])
                row[4] = (wl['date']).strftime('%Y/%m/%d')
                row[5] = 0
                # 
                row[6] = ('')
                # 进度
                row[7] = ('')
                # 修复bugs数量
                row[8] = 0
                # 备注
                row[9] = ('')
                if taskId in self._task2Event.keys() and dt in self._task2Event[taskId] :
                    row[9] = ';'.join(self._task2Event[taskId][dt])
                pass
            
            row = task_workloads[wl['account']][key]
            row[5] += (wl['consumed'])
            # 
            row[6] = ('是' if (wl['work'] or '是'==row[6]) else '否')
            # 进度
            row[7] = ('正常')
            # 修复bugs数量
            #row[8] = ('')
            pass
        #pprint(task_workloads)
        
        # 整理成有序数组
        user2workloads = self._user2workloads
        for k,v in task_workloads.items() :
            user = self._acc2name[k]
            user2workloads[user] = sorted(v.values(),key=lambda z : z[0])
            
            pass
        #pprint(user2workloads)
        
        
        # 
        
    def parseBugFixedData(self):
        '''
        '''
        xls_bugfixed_headers = ['序号','项目','bugID','bug等级','是否按时确认bug','是否按时修复bug','激活次数','是否按格式填写修复说明','备注']
        user2bugfixed = self._d['user2bugfixed'] = {}
        for user,bugsFixed in self._user2BugFixed.items() :
            if user not in user2bugfixed.keys() : user2bugfixed[user] = []
            for bugId,bug in bugsFixed.items() :
                # bug确认时间 ：开发需在bug提交后第二个工作日内确认bug。
                confirmedDate = bug.get('confirmedDate',bug['resolvedDate'])
                activatedDate = bug.get('activatedDate',bug['openedDate'])
                resolvedDate = bug['resolvedDate']
                
                row = ['' for i in range(len(xls_bugfixed_headers))]
                row[0] = 1+len(user2bugfixed[user])
                row[1] = bug['项目']
                row[2] = bugId 
                row[3] = bug['bug等级']
                # 按时确认
                row[4] = '是' if (confirmedDate-activatedDate).days <= 1 else '否'
                row[5] = '是'
                row[6] = bug['activatedCount']
                row[7] = '是' if bug.get('resolvedComment',None) else '否'
                row[8] = ''
                #
                if '否'==row[4] :
                    if bug.get('confirmedDate',None) :
                        row[8] += '实际确认日期:{0};'.format(confirmedDate)
                        pass
                    else :
                        row[8] += '无确认日期;'
                        pass                        
                    pass
                """
                bug修复：
                1级：确认当日内修复bug。
                2级：自确认日起3个工作日（含确认日）内修复bug。
                3,4级：自确认日起一周内（含确认日）修复bug。
                """
                isbugFixExceed = False
                if 1 == bug['bug等级'] :
                    if (resolvedDate-confirmedDate).days > 1 :
                        isbugFixExceed = True
                        pass
                    pass
                elif 2 == bug['bug等级'] :
                    if (resolvedDate-confirmedDate).days > 3 :
                        isbugFixExceed = True
                        pass                    
                    pass
                elif bug['bug等级'] in [3,4]:
                    if (resolvedDate-confirmedDate).days > 7 :
                        isbugFixExceed = True
                        pass                    
                    pass
                #
                if isbugFixExceed :
                    row[5] = '否'
                    row[8] += '实际修复日期:{0};'.format(resolvedDate)
                    pass
                #
                user2bugfixed[user].append(row)
                pass
            pass
        #pprint(user2bugfixed)
        self._user2bugs.update(user2bugfixed)
        #exit(0)
        pass
    def parseBugVerifiedData(self) :
        '''
        '''
        xls_bugverified_headers = ['序号','项目','bugID','是否按时验证bug','是否按格式填写验证结果','备注']
        user2bugverifed = {}
        #pprint(self._user2BugVerified)
        for user,bugs in self._user2BugVerified.items() :
            if user not in user2bugverifed.keys() : user2bugverifed[user] = []
            for actionId,bug in bugs.items():
                row = ['' for i in range(len(xls_bugverified_headers))]
                row[0] = 1 + len(user2bugverifed[user])
                row[1] = bug['项目']
                row[2] = bug['bugId']
                row[3] = '是'
                row[4] = '是' if bug['comment'] else '否'
                row[5] = ''
                if bug['buildDate'] :
                    if bug['verifiedDate'].date() <= bug['buildDate'] :
                        pass
                    else :
                        row[3] = '否'
                        pass
                else :
                    row[5] += '无验证版本信息'
                    pass
                user2bugverifed[user].append(row)
                pass
            pass
        #pprint(user2bugverifed)
        #exit(0)
        self._user2bugs.update(user2bugverifed)
    
    def getWorkloadsData(self):
        '''
        '''
        return self._user2workloads
    
    def getBugsData(self):
        '''
        '''
        return self._user2bugs    
    
def test() :
    base_dir = "d:\\test\\py3"
    #new_report = [base_dir + "\\开发任务检查表 - java.xlsx"]
    new_report = [base_dir + "\\测试任务检查表 .xlsx"]
    print(new_report[0])
    #return
    #             写入
    #wb = Workbook()
    #ws = wb.active
    #ws['A1'] = 35
    #ws.append(["我", "你", "她"])
    # 保存
    #wb.save(filename=new_report[0])
    
    #               读取
    wb1 = load_workbook(filename=new_report[0])
    sheets = wb1.sheetnames  # 获取所有的表格
    for sheet in sheets : print(sheet)
    sheets_first = sheets[-1]    # 获取第一个表
    for sheet in wb1 :
        print(sheet.title,sheet.sheet_state)
    ws1 = wb1[sheets_first]
    
    #print(ws1.)
    #ws1.insert_rows(2,5)
    ws1['B2'] = 'test2'
    ws1['C2'] = 'test2'
    ws1['D2'] = '1000'
    ws1['E2'] = datetime.now().date()
    ws1['F2'] = 3
    ws1['G2'] = '是'
    #print(ws1['G2'].style)
    #row = ws1.row[17]
    print(ws1.sheet_format)
    for s in range(ord('A'),ord('J') ) : 
        ns = chr(s)
        
        ws1['{0}18'.format(ns)].font = copy(ws1['{0}17'.format(ns)].font)
        ws1['{0}18'.format(ns)].fill = copy(ws1['{0}17'.format(ns)].fill)
        ws1['{0}18'.format(ns)].border = copy(ws1['{0}17'.format(ns)].border)
        ws1['{0}18'.format(ns)].number_format = copy(ws1['{0}17'.format(ns)].number_format)
        
        ws1['{0}18'.format(ns)].protection = copy(ws1['{0}17'.format(ns)].protection)
        ws1['{0}18'.format(ns)].alignment = copy(ws1['{0}17'.format(ns)].alignment)
        ws1['{0}18'.format(ns)].style = copy(ws1['{0}17'.format(ns)].style)
        
        ws1['{0}18'.format(ns)].value = copy(ws1['{0}17'.format(ns)].value)
        
    #ws1['B18'] = ws1['B17']
    wb1.save(filename=new_report[0])
    
    pass

def main() :
    '''
    '''
    # 准备
    save_dir = 'D:\\test\\py3'
    template_file = save_dir + '\\TEMP-开发任务检查表.xlsx'
    start_date = date(2018,1,24).strftime('%Y%m%d')
    end_date = date(2018,4,30).strftime('%Y%m%d')    
    print('导出报表根目录 : {0}'.format(save_dir))
    print('模板文件 : {0}'.format(os.path.join(template_file)))
    print('报表周期为 : {0} - {1}'.format(start_date,end_date))
    #
    eh = excelHelper(template_file, save_dir, start_date, end_date)
    eh.scaffold()
    #
    zt = ztHelper()
    zt.collectDataFromDB(start_date, end_date)
    zt.parseData()
    
    workloads = zt.getWorkloadsData()
    bugs = zt.getBugsData()
    #
    eh.writeWorkloads(workloads)
    eh.writeBugs(bugs)
    eh.makeZip()
    
    pass


if __name__ == '__main__':
    main()
    print("--> DONE")
